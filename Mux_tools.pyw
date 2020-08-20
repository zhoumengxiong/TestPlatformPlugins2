# -*- coding: utf-8 -*-
"""
开发者：周梦雄
最后更新日期：2020/8/20
"""
import sys
import os
from PyQt5.QtWidgets import (
    QApplication,
    QMainWindow,
    QMessageBox,
    QTableWidgetItem,
    QAbstractItemView,
)
from Ui_mux_tools import *
from PyQt5.QtCore import QDateTime
import sqlite3
from openpyxl import Workbook
import configparser
import pyodbc
from configuration_databases import *
from datetime import datetime
from PyQt5.QtCore import QRegExp, QThread, pyqtSignal
from PyQt5.QtGui import QRegExpValidator
from urllib.request import urlopen
from urllib.error import URLError


# 重写单相相线程类


class ThreadSingle(QThread):
    signal = pyqtSignal(str)

    def __init__(self):
        super().__init__()

    def run(self):
        ex.session_add_single()
        session_chip_id.commit()
        self.signal.emit("恭喜，上传成功！")


# 重写3105集中器线程类


class ThreadConcentrator3105(QThread):
    signal = pyqtSignal(str)

    def __init__(self):
        super().__init__()

    def run(self):
        ex.session_add_cco3105()
        session_chip_id.commit()
        self.signal.emit("恭喜，上传成功！")


# 重写3911集中器线程类


class ThreadConcentrator3911(QThread):
    signal = pyqtSignal(str)

    def __init__(self):
        super().__init__()

    def run(self):
        ex.session_add_cco3911()
        session_chip_id.commit()
        self.signal.emit("恭喜，上传成功！")


# noinspection SqlDialectInspection
class MyMainWindow(QMainWindow, Ui_STA_database_query):
    def __init__(self):
        super().__init__()
        self.setupUi(self)
        # 审批单号、派工单号加入长度限制
        reg = QRegExp("[0-9]+$")
        approvalValidator = QRegExpValidator(reg)
        self.le_approval_no.setValidator(approvalValidator)
        self.le_approval_no_3105.setValidator(approvalValidator)
        self.le_approval_no_3911.setValidator(approvalValidator)
        # 用来保存ID对应表
        self.result_unique = []
        self.result_unique_3105 = []
        self.result_unique_3911 = []
        # 设置默认产品型态
        self.cb_prod_type_3105.setCurrentIndex(2)
        self.cb_prod_type_3911.setCurrentIndex(2)
        # 查询数据库，设置Qcombobox值
        software_version_all = session.query(
            SoftwareVersion.software_version).all()
        customer_version_all = session.query(
            CustomerVersion.customer_version).all()
        vendor_code_all = session.query(VendorCode.vendor_code).all()
        software_date_all = session.query(SoftwareDate.software_date).all()
        software_version_all_list = [e[0] for e in software_version_all]
        customer_version_all_list = [e[0] for e in customer_version_all]
        vendor_code_all_list = [e[0] for e in vendor_code_all]
        software_date_all_list = [e[0] for e in software_date_all]
        self.cb_version_sw.addItems(software_version_all_list)
        self.cb_vendor_code.addItems(vendor_code_all_list)
        self.cb_date_sw.addItems(software_date_all_list)
        self.cb_ext_version.addItems(customer_version_all_list)

        # 实例化单相线程
        self.thread = ThreadSingle()
        self.thread.signal.connect(
            self.finish)  # 连接线程类中自定义信号槽到本类的自定义槽函数

        # 实例化3105集中器线程
        self.thread_cco3105 = ThreadConcentrator3105()
        self.thread_cco3105.signal.connect(
            self.finish_cco3105)  # 连接线程类中自定义信号槽到本类的自定义槽函数

        # 实例化3911集中器线程
        self.thread_cco3911 = ThreadConcentrator3911()
        self.thread_cco3911.signal.connect(
            self.finish_cco3911)  # 连接线程类中自定义信号槽到本类的自定义槽函数

        # 设置派工单默认前缀
        datetime_now_str = str(datetime.now())
        self.value_order.setText(
            'X' + datetime_now_str[:4] + datetime_now_str[5:7])
        self.value_order_3105.setText(
            'X' + datetime_now_str[:4] + datetime_now_str[5:7])
        self.value_order_3911.setText(
            'X' + datetime_now_str[:4] + datetime_now_str[5:7])
        # 树莓派初始化
        self.tableWidget.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.tableWidget.setColumnCount(3)
        self.tableWidget.setRowCount(99)
        self.tableWidget.resizeColumnsToContents()
        self.tableWidget.resizeRowsToContents()
        self.start_datetime.setDateTime(QDateTime.currentDateTime())
        self.textBrowser.append(
            "注意：3105芯片代码03；3911集中器芯片代码00，STA 01；北京、浙江集中器白名单关闭；"
        )
        self.textBrowser.setStyleSheet("* { color: #0000FF;}")
        self.statusbar.setStyleSheet(
            "* { color: #FF6666;font-size:30px;font-weight:bold;}"
        )
        self.listWidget.setSpacing(7)
        # 数据库路径
        # 生产路径
        # db_file = r"C:\Users\Lenovo\Desktop\更新带屏读ID软件\最新读ID(带屏)Debug  20190621\Debug带瓶的\MyProtocol.db"
        # 测试路径
        db_file = r"MyProtocol.db"
        # 创建数据库连接对象
        self.conn = sqlite3.connect(db_file)
        # 创建游标对象
        self.cur = self.conn.cursor()
        self.start_date = self.start_datetime.dateTime().toString("yyyy-MM-dd HH:mm")
        self.sqlstring = r"SELECT ChipID,ModID,TTime FROM DataBackUp where ChipID<>'' and TTime>=? order by TTime asc;"
        self.query.clicked.connect(self.click_query)
        self.start_datetime.dateTimeChanged.connect(self.on_datetime_changed)
        self.export_id.clicked.connect(self.export_id_to_excel)
        self.btn_configure.clicked.connect(self.nv_query)
        self.bt_save.clicked.connect(self.write_ini)

        # 3105 CCO初始化
        self.tableWidget_3105.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.tableWidget_3105.setColumnCount(4)
        self.tableWidget_3105.setRowCount(99)
        self.tableWidget_3105.resizeColumnsToContents()
        self.tableWidget_3105.resizeRowsToContents()
        self.value_start_datetime_3105.setDateTime(QDateTime.currentDateTime())
        self.textBrowser_3105.append(
            "注意：3105芯片代码03；3911集中器芯片代码00，STA 01；北京、浙江集中器白名单关闭；"
        )
        self.textBrowser_3105.setStyleSheet("* { color: #0000FF;}")
        self.statusbar.setStyleSheet(
            "* { color: #FF6666;font-size:30px;font-weight:bold;}"
        )
        # 数据库驱动
        # sql_driver_3105 = r'DSN=芯片ID;DBQ=C:\PC_PRODCHECK\DATA\Equip_sta.mdb;DefaultDir=C:\PC_PRODCHECK\DATA;DriverId=25;FIL=MS Access;MaxBufferSize=2048;PageTimeout=5;'
        self.sql_driver_3105 = r"DRIVER=Microsoft Access Driver (*.mdb, *.accdb);DBQ=C:/HiStudio-工装/3105集中器通信模块出厂检测/data/ndm/NoStaEquip/NoStaCCOCheckDB.mdb"
        # 创建数据库连接对象
        """self.conn_3105 = pyodbc.connect(self.sql_driver_3105)
        # 创建游标对象
        self.cur_3105 = self.conn_3105.cursor()"""
        self.start_date_3105 = self.value_start_datetime_3105.dateTime().toString(
            "yyyy-MM-dd HH:mm")
        self.sqlstring_3105 = r"SELECT 总体测试结果,芯片ID值,模块ID值,日期 FROM NoStaTableCCOCheck where 芯片ID值<>'' and 日期>=? order by 日期 asc;"
        self.btn_id_query_3105.clicked.connect(self.click_query_3105)
        self.value_start_datetime_3105.dateTimeChanged.connect(
            self.on_datetime_changed_3105)
        self.btn_export_id_3105.clicked.connect(self.export_id_to_excel_3105)

        # 3911 CCO初始化
        self.tableWidget_3911.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.tableWidget_3911.setColumnCount(4)
        self.tableWidget_3911.setRowCount(99)
        self.tableWidget_3911.resizeColumnsToContents()
        self.tableWidget_3911.resizeRowsToContents()
        self.value_start_datetime_3911.setDateTime(QDateTime.currentDateTime())
        self.textBrowser_3911.append(
            "注意：3105芯片代码03；3911集中器芯片代码00，STA 01；北京、浙江集中器白名单关闭；"
        )
        self.textBrowser_3911.setStyleSheet("* { color: #0000FF;}")
        self.statusbar.setStyleSheet(
            "* { color: #FF6666;font-size:30px;font-weight:bold;}"
        )
        # 数据库驱动
        # sql_driver_3911 = r'DSN=芯片ID;DBQ=C:\PC_PRODCHECK\DATA\Equip_sta.mdb;DefaultDir=C:\PC_PRODCHECK\DATA;DriverId=25;FIL=MS Access;MaxBufferSize=2048;PageTimeout=5;'
        self.sql_driver_3911 = r"DRIVER=Microsoft Access Driver (*.mdb, *.accdb);DBQ=C:/HiStudio-工装/集中器通信模块出厂检测-3911/data/ndm/NoStaEquip/NoStaCCOCheckDB.mdb"
        # 创建数据库连接对象
        """self.conn_3911 = pyodbc.connect(self.sql_driver_3911)
        # 创建游标对象
        self.cur_3911 = self.conn_3911.cursor()"""
        self.start_date_3911 = self.value_start_datetime_3911.dateTime().toString(
            "yyyy-MM-dd HH:mm")
        self.sqlstring_3911 = r"SELECT 总体测试结果,芯片ID值,模块ID值,日期 FROM NoStaTableCCOCheck where 芯片ID值<>'' and 日期>=? order by 日期 asc;"
        self.btn_id_query_3911.clicked.connect(self.click_query_3911)
        self.value_start_datetime_3911.dateTimeChanged.connect(
            self.on_datetime_changed_3911)
        self.btn_export_id_3911.clicked.connect(self.export_id_to_excel_3911)
        self.btn_software_version.clicked.connect(self.add_software_version)
        self.btn_customer_version.clicked.connect(self.add_customer_version)
        self.btn_vendor_code.clicked.connect(self.add_vendor_code)
        self.btn_software_date.clicked.connect(self.add_date)
        self.btn_upload.clicked.connect(self.upload)
        self.btn_upload_3105.clicked.connect(self.upload_3105)
        self.btn_upload_3911.clicked.connect(self.upload_3911)
        self.le_approval_no.editingFinished.connect(self.approval_check)
        self.le_approval_no_3105.editingFinished.connect(self.approval_check_3105)
        self.le_approval_no_3911.editingFinished.connect(self.approval_check_3911)

        self.show()

    # 检查单相是否输入6位审批单号
    def approval_check(self):
        if len(self.le_approval_no.text()) != 6:
            QMessageBox.warning(
                self, '错误：', '您输入的审批单号不是6位，请重新输入！', QMessageBox.Ok)
            self.le_approval_no.setFocus()

    # 检查3105集中器是否输入6位审批单号
    def approval_check_3105(self):
        if len(self.le_approval_no_3105.text()) != 6:
            QMessageBox.warning(
                self, '错误：', '您输入的审批单号不是6位，请重新输入！', QMessageBox.Ok)
            self.le_approval_no_3105.setFocus()

    # 检查3911集中器是否输入6位审批单号
    def approval_check_3911(self):
        if len(self.le_approval_no_3911.text()) != 6:
            QMessageBox.warning(
                self, '错误：', '您输入的审批单号不是6位，请重新输入！', QMessageBox.Ok)
            self.le_approval_no_3911.setFocus()

    def on_datetime_changed(self):
        self.start_date = self.start_datetime.dateTime().toString("yyyy-MM-dd HH:mm")

    def click_query(self):
        self.tableWidget.clearContents()  # 每一次查询时清除表格中信息
        # 执行查询（传递开始测试日期时间参数）
        self.cur.execute(self.sqlstring, (self.start_date,))
        result_temp = self.cur.fetchall()
        result_id = [(r[0], r[1]) for r in result_temp]
        for i in result_id:
            if i not in self.result_unique:
                self.result_unique.append(i)
        try:
            # 自动设置ID倒数5个字符
            self.lineEdit.setText(result_temp[0][0][-5:])
            for k, i in enumerate(result_temp):
                print("----------", i)
                for w, j in enumerate(i):
                    if type(j) != str:
                        newItem = QTableWidgetItem(str(j))
                    else:
                        newItem = QTableWidgetItem(j)
                    # 根据循环标签一次对table中的格子进行设置
                    self.tableWidget.setItem(k, w, newItem)
            self.tableWidget.resizeColumnsToContents()
            self.tableWidget.resizeRowsToContents()
            self.textBrowser.setText("")
            self.textBrowser.append(
                "SELECT ChipID,ModID,TTime FROM DataBackUp where ChipID<>'' and TTime>=%r order by TTime asc;"
                % self.start_date
            )
            print("find button pressed")
        except IndexError:
            self.result_unique = []
            QMessageBox.warning(self, '提示：', '查询结果为空！', QMessageBox.Ok)

    def export_id_to_excel(self):
        if not self.result_unique:
            QMessageBox.warning(self, '提示：', '请先点击查询按钮确认ID对应表有无异常！', QMessageBox.Ok)
            return
        if len(self.value_order.text().strip()) <= 10:
            QMessageBox.warning(
                self, '错误：', '派工单号输入有误，请重新输入！', QMessageBox.Ok)
            self.value_order.setFocus()
            return
        else:
            wo = self.value_order.text().strip().upper()
            if self.cb_prod_type.currentText() == "通信单元（集中器I型/HPLC）":
                wo1 = wo + "-" + self.cb_prod_type.currentText().split('/')[0][-5:] + ".xlsx"
            elif self.cb_prod_type.currentText() == "II型采集器（HPLC）":
                wo1 = wo + "-" + self.cb_prod_type.currentText().split('（')[0] + ".xlsx"
            else:
                wo1 = wo + "-" + self.cb_prod_type.currentText().split('/')[0][-2:] + ".xlsx"
            # 生产环境工作簿保存路径
            # path_name = os.path.join(
            #     r"C:\Users\Lenovo\Desktop\ID清单，请手下留情，勿删！！！", wo1)
            # 测试环境工作部保存路径
            path_name = os.path.join(os.getcwd(), wo1)
            # 新建工作簿
            wb = Workbook(path_name)
            ws = wb.create_sheet(wo, 0)
            ws.append(["芯片ID", "模块ID"])
            for row in self.result_unique:
                ws.append(list(row))
            self.statusbar.showMessage(
                "本批测试 %s 个模块，请注意检查是否有漏测！" % len(self.result_unique), 100000)
            if self.result_unique[0][0][-5:] != self.lineEdit.text().upper():
                self.statusbar.clearMessage()
                QMessageBox.warning(
                    self, "警告！", "你的首个ID不正确，请排查原因！", QMessageBox.Ok)
            else:
                wb.save(path_name)
                QMessageBox.information(
                    self, "好消息！", "ID对应表已成功导出到excel表格！请核对左下角状态栏信息！", QMessageBox.Ok
                )

    def nv_query(self):
        # 初始化
        conf = configparser.ConfigParser()
        # 配置文件的绝对路径
        # 生产路径
        # conf_path = r"C:\Users\Lenovo\Desktop\更新带屏读ID软件\最新读ID(带屏)Debug  20190621\Debug带瓶的\IniFile\FiterParam.ini"
        # 测试路径
        conf_path = r"FiterParam.ini"
        # 读取配置文件
        conf.read(conf_path)
        # 返回section中option的值
        nv_configure = "软件版本：%s 芯片代码：%s 版本日期：%s 外部版本：%s 厂商代码：%s" % (
            conf.get("ErJiBiDui", "Value1"), conf.get("ErJiBiDui", "Value2"),
            conf.get("ErJiBiDui", "Value3"), conf.get("ErJiBiDui", "Value4"),
            conf.get("ErJiBiDui", "Value5"))
        self.textBrowser.setText("")
        self.textBrowser.append(nv_configure)

    def write_ini(self):
        config = configparser.ConfigParser()
        # 生产路径
        # path_name = r"C:\Users\Lenovo\Desktop\更新带屏读ID软件\最新读ID(带屏)Debug  20190621\Debug带瓶的\IniFile\FiterParam.ini"
        # 测试路径
        path_name = r"FiterParam.ini"
        config.read(path_name)  # 读文件
        section = r"ErJiBiDui"
        # 新增/修改配置文件的键值
        if len(self.cb_version_sw.currentText()) == 14:
            if self.cb_type.currentText() == '单相表':
                self.cb_prod_type.setCurrentIndex(0)
                config.set(section, 'Value1',
                           self.cb_version_sw.currentText()[0:11] + '100')
            else:
                self.cb_prod_type.setCurrentIndex(2)
                config.set(section, 'Value1',
                           self.cb_version_sw.currentText()[0:11] + '000')
        else:
            config.set(section, 'Value1', self.cb_version_sw.currentText())
        config.set(section, 'Value2', self.cb_chipcode.currentText())
        config.set(section, 'Value3', self.cb_date_sw.currentText())
        config.set(section, 'Value4', (self.cb_ext_version.currentText()[
                                       2:] + self.cb_ext_version.currentText()[:2]))
        config.set(section, 'Value5', self.cb_vendor_code.currentText())
        with open(path_name, 'w') as configfile:
            config.write(configfile)
        self.statusbar.setStyleSheet(
            "* { color: #00CD00;font-size:30px;font-weight:bold;}")
        self.statusbar.showMessage("配置文件修改成功！", 3000)

    def closeEvent(self, event):  # 函数名固定不可变
        reply = QtWidgets.QMessageBox.question(
            self, u'警告', u'确认退出?', QtWidgets.QMessageBox.Yes, QtWidgets.QMessageBox.No)
        # QtWidgets.QMessageBox.question(self,u'弹窗名',u'弹窗内容',选项1,选项2)
        if reply == QtWidgets.QMessageBox.Yes:
            session_chip_id.close()
            session.close()
            self.conn.commit()
            self.cur.close()
            self.conn.close()
            self.conn_3105.commit()
            self.cur_3105.close()
            self.conn_3105.close()
            self.conn_3911.commit()
            self.cur_3911.close()
            self.conn_3911.close()
            event.accept()  # 关闭窗口
        else:
            event.ignore()  # 忽视点击X事件

    def add_software_version(self):
        if self.le_software_version.text() == '':
            self.statusbar.clearMessage()
            QMessageBox.warning(
                self, "警告！", "软件版本不能为空！", QMessageBox.Ok)
        elif session.query(SoftwareVersion.software_version).filter_by(
                software_version=self.le_software_version.text().strip().upper()).first():
            self.statusbar.clearMessage()
            QMessageBox.warning(
                self, "警告！", "数据库中已存在该数据，请勿重复添加！", QMessageBox.Ok)
        else:
            self.cb_version_sw.addItem(
                self.le_software_version.text().strip().upper())
            new_software_version = SoftwareVersion(
                software_version=self.le_software_version.text().strip().upper())
            session.add(new_software_version)
            session.commit()
            self.statusbar.setStyleSheet(
                "* { color: #00CD00;font-size:30px;font-weight:bold;}")
            self.statusbar.showMessage("软件版本添加成功！", 3000)

    def add_customer_version(self):
        if self.le_customer_version.text() == '':
            self.statusbar.clearMessage()
            QMessageBox.warning(
                self, "警告！", "外部版本不能为空！", QMessageBox.Ok)
        elif session.query(CustomerVersion.customer_version).filter_by(
                customer_version=self.le_customer_version.text().strip()).first():
            self.statusbar.clearMessage()
            QMessageBox.warning(
                self, "警告！", "数据库中已存在该数据，请勿重复添加！", QMessageBox.Ok)
        else:
            self.cb_ext_version.addItem(
                self.le_customer_version.text().strip())
            new_customer_version = CustomerVersion(
                customer_version=self.le_customer_version.text().strip())
            session.add(new_customer_version)
            session.commit()
            self.statusbar.setStyleSheet(
                "* { color: #00CD00;font-size:30px;font-weight:bold;}")
            self.statusbar.showMessage("外部版本添加成功！", 3000)

    def add_vendor_code(self):
        if self.le_vendor_code.text() == '':
            self.statusbar.clearMessage()
            QMessageBox.warning(
                self, "警告！", "厂商代码不能为空！", QMessageBox.Ok)
        elif session.query(VendorCode.vendor_code).filter_by(
                vendor_code=self.le_vendor_code.text().strip().upper()).first():
            self.statusbar.clearMessage()
            QMessageBox.warning(
                self, "警告！", "数据库中已存在该数据，请勿重复添加！", QMessageBox.Ok)
        else:
            self.cb_vendor_code.addItem(
                self.le_vendor_code.text().strip().upper())
            new_vendor_code = VendorCode(
                vendor_code=self.le_vendor_code.text().strip().upper())
            session.add(new_vendor_code)
            session.commit()
            self.statusbar.setStyleSheet(
                "* { color: #00CD00;font-size:30px;font-weight:bold;}")
            self.statusbar.showMessage("厂商代码添加成功！", 3000)

    def add_date(self):
        if self.le_date.text() == '':
            self.statusbar.clearMessage()
            QMessageBox.warning(
                self, "警告！", "版本日期不能为空！", QMessageBox.Ok)
        elif session.query(SoftwareDate.software_date).filter_by(software_date=self.le_date.text().strip()).first():
            self.statusbar.clearMessage()
            QMessageBox.warning(
                self, "警告！", "数据库中已存在该数据，请勿重复添加！", QMessageBox.Ok)
        else:
            self.cb_date_sw.addItem(self.le_date.text().strip())
            new_software_date = SoftwareDate(
                software_date=self.le_date.text().strip())
            session.add(new_software_date)
            session.commit()
            self.statusbar.setStyleSheet(
                "* { color: #00CD00;font-size:30px;font-weight:bold;}")
            self.statusbar.showMessage("版本日期添加成功！", 3000)

    def on_datetime_changed_3105(self):
        self.start_date_3105 = self.value_start_datetime_3105.dateTime().toString(
            "yyyy-MM-dd HH:mm")

    def click_query_3105(self):
        self.tableWidget_3105.clearContents()  # 每一次查询时清除表格中信息
        # 执行查询（传递开始测试日期时间参数）
        self.cur_3105.execute(self.sqlstring_3105, self.start_date_3105)
        result_temp = self.cur_3105.fetchall()
        result_id = [(r[1], r[2]) for r in result_temp]
        for i in result_id:
            if i not in self.result_unique_3105:
                self.result_unique_3105.append(i)
        try:
            # 自动设置ID倒数5个字符
            self.value_id_3105.setText(result_temp[0][1][-5:])
            for k, i in enumerate(result_temp):
                print("----------", i)
                for w, j in enumerate(i):
                    if type(j) != str:
                        newItem = QTableWidgetItem(str(j))
                    else:
                        newItem = QTableWidgetItem(j)
                    # 根据循环标签一次对table中的格子进行设置
                    self.tableWidget_3105.setItem(k, w, newItem)
            self.tableWidget_3105.resizeColumnsToContents()
            self.tableWidget_3105.resizeRowsToContents()
            self.textBrowser_3105.setText("")
            self.textBrowser_3105.append(
                "SELECT 总体测试结果,芯片ID值,模块ID值,日期 FROM NoStaTableCCOCheck where 芯片ID值<>'' and 日期>=%r order by 日期 asc;"
                % self.start_date_3105
            )
            print("find button pressed")
        except IndexError:
            self.result_unique_3105 = []
            QMessageBox.warning(self, '提示：', '查询结果为空！', QMessageBox.Ok)

    def export_id_to_excel_3105(self):
        if not self.result_unique_3105:
            QMessageBox.warning(self, '提示：', '请先点击查询按钮确认ID对应表有无异常！', QMessageBox.Ok)
            return
        if len(self.value_order_3105.text().strip()) <= 10:
            QMessageBox.warning(
                self, '错误：', '派工单号输入有误，请重新输入！', QMessageBox.Ok)
            self.value_order_3105.setFocus()
            return
        else:
            wo = self.value_order_3105.text().strip().upper()
            if self.cb_prod_type_3105.currentText() == "通信单元（集中器I型/HPLC）":
                wo1 = wo + "-" + self.cb_prod_type_3105.currentText().split('/')[0][-5:] + ".xlsx"
            elif self.cb_prod_type_3105.currentText() == "II型采集器（HPLC）":
                wo1 = wo + "-" + self.cb_prod_type_3105.currentText().split('（')[0] + ".xlsx"
            else:
                wo1 = wo + "-" + self.cb_prod_type_3105.currentText().split('/')[0][-2:] + ".xlsx"
            # 生产环境工作簿保存路径
            # path_name = os.path.join(
            #     r"C:\Users\Lenovo\Desktop\ID清单，请手下留情，勿删！！！", wo1)
            # 生产环境工作簿保存路径
            # path_name = os.path.join(
            #     r"C:\Users\Lenovo\Desktop\ID清单，请手下留情，勿删！！！", wo1)
            # 测试环境工作簿保存路径
            path_name = os.path.join(os.getcwd(), wo1)
            # 新建工作簿
            wb = Workbook(path_name)
            ws = wb.create_sheet(wo, 0)
            ws.append(["芯片ID", "模块ID"])
            for row in self.result_unique_3105:
                ws.append(list(row))
            self.statusbar.showMessage(
                "本批测试 %s 个模块，请注意检查是否有漏测！" % len(self.result_unique_3105), 100000)
            if self.result_unique_3105[0][0][-5:] != self.value_id_3105.text().upper():
                self.statusbar.clearMessage()
                QMessageBox.warning(
                    self, "警告！", "你的首个ID不正确，请排查原因！", QMessageBox.Ok)
            else:
                wb.save(path_name)
                QMessageBox.information(
                    self, "好消息！", "ID对应表已成功导出到excel表格！请核对左下角状态栏信息！", QMessageBox.Ok
                )

    def on_datetime_changed_3911(self):
        self.start_date_3911 = self.value_start_datetime_3911.dateTime().toString(
            "yyyy-MM-dd HH:mm")

    def click_query_3911(self):
        self.tableWidget_3911.clearContents()  # 每一次查询时清除表格中信息
        # 查询结果
        self.cur_3911.execute(self.sqlstring_3911, self.start_date_3911)
        result_temp = self.cur_3911.fetchall()
        result_id = [(r[1], r[2]) for r in result_temp]
        for i in result_id:
            if i not in self.result_unique_3911:
                self.result_unique_3911.append(i)
        try:
            # 自动设置ID倒数5个字符
            self.value_id_3911.setText(result_temp[0][1][-5:])
            for k, i in enumerate(result_temp):
                print("----------", i)
                for w, j in enumerate(i):
                    if type(j) != str:
                        newItem = QTableWidgetItem(str(j))
                    else:
                        newItem = QTableWidgetItem(j)
                    # 根据循环标签一次对table中的格子进行设置
                    self.tableWidget_3911.setItem(k, w, newItem)
            self.tableWidget_3911.resizeColumnsToContents()
            self.tableWidget_3911.resizeRowsToContents()
            self.textBrowser_3911.setText("")
            self.textBrowser_3911.append(
                "SELECT 总体测试结果,芯片ID值,模块ID值,日期 FROM NoStaTableCCOCheck where 芯片ID值<>'' and 日期>=%r order by 日期 asc;"
                % self.start_date_3911
            )
            print("find button pressed")
        except IndexError:
            self.result_unique_3911 = []
            QMessageBox.warning(self, '提示：', '查询结果为空！', QMessageBox.Ok)

    def export_id_to_excel_3911(self):
        if not self.result_unique_3911:
            QMessageBox.warning(self, '提示：', '请先点击查询按钮确认ID对应表有无异常！', QMessageBox.Ok)
            return
        if len(self.value_order_3911.text().strip()) <= 10:
            QMessageBox.warning(
                self, '错误：', '派工单号输入有误，请重新输入！', QMessageBox.Ok)
            self.value_order_3911.setFocus()
            return
        else:
            wo = self.value_order_3911.text().strip().upper()
            if self.cb_prod_type_3911.currentText() == "通信单元（集中器I型/HPLC）":
                wo1 = wo + "-" + self.cb_prod_type_3911.currentText().split('/')[0][-5:] + ".xlsx"
            elif self.cb_prod_type_3911.currentText() == "II型采集器（HPLC）":
                wo1 = wo + "-" + self.cb_prod_type_3911.currentText().split('（')[0] + ".xlsx"
            else:
                wo1 = wo + "-" + self.cb_prod_type_3911.currentText().split('/')[0][-2:] + ".xlsx"
            # 生产环境工作簿保存路径
            # path_name = os.path.join(
            #     r"C:\Users\Lenovo\Desktop\ID清单，请手下留情，勿删！！！", wo1)
            # 测试环境工作簿保存路径
            path_name = os.path.join(os.getcwd(), wo1)
            # 新建工作簿
            wb = Workbook(path_name)
            ws = wb.create_sheet(wo, 0)
            ws.append(["芯片ID", "模块ID"])
            for row in self.result_unique_3911:
                ws.append(list(row))
            self.statusbar.showMessage(
                "本批测试 %s 个模块，请注意检查是否有漏测！" % len(self.result_unique_3911))
            if self.result_unique_3911[0][0][-5:] != self.value_id_3911.text().upper():
                self.statusbar.clearMessage()
                QMessageBox.warning(
                    self, "警告！", "你的首个ID不正确，请排查原因！", QMessageBox.Ok)
            else:
                wb.save(path_name)
                QMessageBox.information(
                    self, "好消息！", "ID对应表已成功导出到excel表格！请核对左下角状态栏信息！", QMessageBox.Ok
                )

    def upload(self):
        if len(self.le_approval_no.text()) == 0:
            QMessageBox.warning(
                self, '错误：', '您未输入审批单号，请输入！', QMessageBox.Ok)
            self.le_approval_no.setFocus()
            return
        if len(self.value_order.text().strip()) <= 10:
            QMessageBox.warning(
                self, '错误：', '派工单号输入有误，请重新输入！', QMessageBox.Ok)
            self.value_order.setFocus()
            return
        if not self.result_unique:
            QMessageBox.warning(
                self, '提示：', '请先点击查询按钮确认ID对应表有无异常！', QMessageBox.Ok)
        else:
            try:
                # 判定服务器连接状态
                urlopen("https://www.baidu.com")
            except URLError:
                QMessageBox.warning(
                    self, '警告：', '无法连接到数据库，请检查网络连接状态！', QMessageBox.Ok)
            else:
                self.work()

    def upload_3105(self):
        if len(self.le_approval_no_3105.text()) == 0:
            QMessageBox.warning(
                self, '错误：', '您未输入审批单号，请输入！', QMessageBox.Ok)
            self.le_approval_no_3105.setFocus()
            return
        if len(self.value_order_3105.text().strip()) <= 10:
            QMessageBox.warning(
                self, '错误：', '派工单号输入有误，请重新输入！', QMessageBox.Ok)
            self.value_order_3105.setFocus()
            return
        if not self.result_unique_3105:
            QMessageBox.warning(
                self, '提示：', '请先点击查询按钮确认ID对应表有无异常！', QMessageBox.Ok)
        else:
            try:
                # 判定服务器连接状态
                urlopen("https://www.baidu.com")
            except URLError:
                QMessageBox.warning(
                    self, '警告：', '无法连接到数据库，请检查网络连接状态！', QMessageBox.Ok)
            else:
                self.work_cco3105()

    def upload_3911(self):
        if len(self.le_approval_no_3911.text()) == 0:
            QMessageBox.warning(
                self, '错误：', '您未输入审批单号，请输入！', QMessageBox.Ok)
            self.le_approval_no_3911.setFocus()
            return
        if len(self.value_order_3911.text().strip()) <= 10:
            QMessageBox.warning(
                self, '错误：', '派工单号输入有误，请重新输入！', QMessageBox.Ok)
            self.value_order_3911.setFocus()
            return
        if not self.result_unique_3911:
            QMessageBox.warning(
                self, '提示：', '请先点击查询按钮确认ID对应表有无异常！', QMessageBox.Ok)
        else:
            try:
                # 判定服务器连接状态
                urlopen("https://www.baidu.com")
            except URLError:
                QMessageBox.warning(
                    self, '警告：', '无法连接到数据库，请检查网络连接状态！', QMessageBox.Ok)
            else:
                self.work_cco3911()

    # 定义单相线程启动函数
    def work(self):
        self.thread.start()
        self.btn_upload.setDisabled(True)

    # 定义3105集中器线程启动函数
    def work_cco3105(self):
        self.thread_cco3105.start()
        self.btn_upload_3105.setDisabled(True)

    # 定义3911集中器线程启动函数
    def work_cco3911(self):
        self.thread_cco3911.start()
        self.btn_upload_3911.setDisabled(True)

    # 定义单相线程完成后槽函数
    def finish(self, info):
        self.statusbar.setStyleSheet(
            "* { color: #00CD00;font-size:30px;font-weight:bold;}")
        self.statusbar.showMessage(info, 3000)
        self.btn_upload.setEnabled(True)

    # 定义3105集中器线程完成后槽函数
    def finish_cco3105(self, info):
        self.statusbar.setStyleSheet(
            "* { color: #00CD00;font-size:30px;font-weight:bold;}")
        self.statusbar.showMessage(info, 3000)
        self.btn_upload_3105.setEnabled(True)

    # 定义3911集中器线程完成后槽函数
    def finish_cco3911(self, info):
        self.statusbar.setStyleSheet(
            "* { color: #00CD00;font-size:30px;font-weight:bold;}")
        self.statusbar.showMessage(info, 3000)
        self.btn_upload_3911.setEnabled(True)

    # 单相chip_id表数据添加到会话
    def session_add(self):
        for tup in self.result_unique:
            if tup[1] and not session_chip_id.query(ChipId).filter_by(
                    chip_id=tup[0]).first() and not session_chip_id.query(ChipId).filter_by(asset_no=tup[1]).first():
                new_work_order_no = WorkOrderNo(
                    work_order_no=self.value_order.text().strip().upper())
                new_approval_no = ApprovalNo(
                    approval_no=self.le_approval_no.text())
                new_product_category = ProductCategory(
                    product_category=self.cb_prod_type.currentText())
                if not session_chip_id.query(WorkOrderNo).filter_by(
                        work_order_no=self.value_order.text().strip().upper()).first():
                    session_chip_id.add(new_work_order_no)
                    session_chip_id.commit()
                if not session_chip_id.query(ApprovalNo).filter_by(
                        approval_no=self.le_approval_no.text()).first():
                    session_chip_id.add(new_approval_no)
                    session_chip_id.commit()
                if not session_chip_id.query(ProductCategory).filter_by(
                        product_category=self.cb_prod_type.currentText()).first():
                    session_chip_id.add(new_product_category)
                    session_chip_id.commit()
                row = ChipId(chip_id=tup[0], asset_no=tup[1],
                             work_order_no_id=session_chip_id.query(WorkOrderNo).filter_by(
                                 work_order_no=self.value_order.text().strip().upper()).first().id,
                             approval_no_id=session_chip_id.query(ApprovalNo).filter_by(
                                 approval_no=self.le_approval_no.text()).first().id,
                             product_category_id=session_chip_id.query(ProductCategory).filter_by(
                                 product_category=self.cb_prod_type.currentText()).first().id)
                session_chip_id.add(row)

    # 3105集中器chip_id表数据添加到会话
    def session_add_cco3105(self):
        for tup in self.result_unique_3105:
            if tup[1] and not session_chip_id.query(ChipId).filter_by(
                    chip_id=tup[0]).first() and not session_chip_id.query(ChipId).filter_by(asset_no=tup[1]).first():
                new_work_order_no = WorkOrderNo(
                    work_order_no=self.value_order_3105.text().strip().upper())
                new_approval_no = ApprovalNo(
                    approval_no=self.le_approval_no_3105.text())
                new_product_category = ProductCategory(
                    product_category=self.cb_prod_type_3105.currentText())
                if not session_chip_id.query(WorkOrderNo).filter_by(
                        work_order_no=self.value_order_3105.text().strip().upper()).first():
                    session_chip_id.add(new_work_order_no)
                    session_chip_id.commit()
                if not session_chip_id.query(ApprovalNo).filter_by(
                        approval_no=self.le_approval_no_3105.text()).first():
                    session_chip_id.add(new_approval_no)
                    session_chip_id.commit()
                if not session_chip_id.query(ProductCategory).filter_by(
                        product_category=self.cb_prod_type_3105.currentText()).first():
                    session_chip_id.add(new_product_category)
                    session_chip_id.commit()
                row = ChipId(chip_id=tup[0], asset_no=tup[1],
                             work_order_no_id=session_chip_id.query(WorkOrderNo).filter_by(
                                 work_order_no=self.value_order_3105.text().strip().upper()).first().id,
                             approval_no_id=session_chip_id.query(ApprovalNo).filter_by(
                                 approval_no=self.le_approval_no_3105.text()).first().id,
                             product_category_id=session_chip_id.query(ProductCategory).filter_by(
                                 product_category=self.cb_prod_type_3105.currentText()).first().id)
                session_chip_id.add(row)

    # 3911集中器chip_id表数据添加到会话
    def session_add_cco3911(self):
        for tup in self.result_unique_3911:
            if tup[1] and not session_chip_id.query(ChipId).filter_by(
                    chip_id=tup[0]).first() and not session_chip_id.query(ChipId).filter_by(asset_no=tup[1]).first():
                new_work_order_no = WorkOrderNo(
                    work_order_no=self.value_order_3911.text().strip().upper())
                new_approval_no = ApprovalNo(
                    approval_no=self.le_approval_no_3911.text())
                new_product_category = ProductCategory(
                    product_category=self.cb_prod_type_3911.currentText())
                if not session_chip_id.query(WorkOrderNo).filter_by(
                        work_order_no=self.value_order_3911.text().strip().upper()).first():
                    session_chip_id.add(new_work_order_no)
                    session_chip_id.commit()
                if not session_chip_id.query(ApprovalNo).filter_by(
                        approval_no=self.le_approval_no_3911.text()).first():
                    session_chip_id.add(new_approval_no)
                    session_chip_id.commit()
                if not session_chip_id.query(ProductCategory).filter_by(
                        product_category=self.cb_prod_type_3911.currentText()).first():
                    session_chip_id.add(new_product_category)
                    session_chip_id.commit()
                row = ChipId(chip_id=tup[0], asset_no=tup[1],
                             work_order_no_id=session_chip_id.query(WorkOrderNo).filter_by(
                                 work_order_no=self.value_order_3911.text().strip().upper()).first().id,
                             approval_no_id=session_chip_id.query(ApprovalNo).filter_by(
                                 approval_no=self.le_approval_no_3911.text()).first().id,
                             product_category_id=session_chip_id.query(ProductCategory).filter_by(
                                 product_category=self.cb_prod_type_3911.currentText()).first().id)
                session_chip_id.add(row)


if __name__ == "__main__":
    app = QApplication(sys.argv)
    ex = MyMainWindow()
    sys.exit(app.exec_())
